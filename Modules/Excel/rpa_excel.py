from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import pandas as pd
from pandas.api.types import CategoricalDtype
from openpyxl.utils.dataframe import dataframe_to_rows

class RPAExcel:
    def __init__(self, config_data, rpa_tracker):
        self.data = config_data
        self.rpa_tracker = rpa_tracker
        self.file_path = None
        self.current_wb = None
        self.active_ws = None

    def open_excel(self, file_path):
        try:
            if not file_path:
                raise Exception("File path cannot be null")

            self.file_path = file_path
            self.current_wb = load_workbook(self.file_path)
            self.active_ws = self.current_wb.sheetnames[0]
        except Exception as e:
            print(f"Error opening Excel file: {str(e)}")

    def fill_data(self, data, sheet_name=None, start_row=0, start_col=0, index=False):
        if sheet_name:
            self.set_worksheet(sheet_name)

        df = pd.DataFrame(data)
        try:
            with pd.ExcelWriter(self.file_path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name=self.active_ws, index=index, header=False, startrow=start_row, startcol=start_col)
        except Exception as e:
            print(f"Error filling data: {str(e)}")

    def insert_image(self, image_path, cell, sheet_name=None):
        try:
            if not self.current_wb:
                self.current_wb = load_workbook(self.file_path)
                
            ws = self.current_wb[sheet_name] if sheet_name else self.current_wb[self.active_ws]
            img = Image(image_path)
            ws.add_image(img, cell)
            self.current_wb.save(self.file_path)
        except Exception as e:
            print(f"Error inserting image: {str(e)}")

    def set_format_cells(self, sheet_name, cell_list, format_type):
        try:
            if not self.current_wb:
                self.current_wb = load_workbook(self.file_path)

            ws = self.current_wb[sheet_name]
            format_map = {
                "string": "@", "number": "0", "currency": "$#,##0.00", 
                "date": "YYYY-MM-DD", "percentage": "0.00%"
            }

            for cell_ref in cell_list:
                cell = ws[cell_ref]
                cell.number_format = format_map.get(format_type, "@")
            self.current_wb.save(self.file_path)
        except Exception as e:
            print(f"Error setting format: {str(e)}")

    def insert_formula(self, sheet_name, list_column, formula_type, direction="vertical"):
        try:
            if not self.current_wb:
                self.current_wb = load_workbook(self.file_path)
            
            ws = self.current_wb[sheet_name]
            formula_map = {"sum": "SUM", "average": "AVERAGE", "count": "COUNT", "max": "MAX", "min": "MIN"}
            formula = formula_map.get(formula_type)
            
            if direction == "vertical":
                for col in list_column:
                    col_idx = ord(col) - 64
                    ws.cell(row=ws.max_row + 1, column=col_idx, value=f"={formula}({col}1:{col}{ws.max_row})")
            elif direction == "horizontal":
                for row in range(1, ws.max_row+1):
                    start_col = list_column[0]
                    end_col = list_column[-1]
                    ws.cell(row=row, column=ord(end_col) - 63, value=f"={formula}({start_col}{row}:{end_col}{row})")
            
            self.current_wb.save(self.file_path)
        except Exception as e:
            print(f"Error inserting formula: {str(e)}")

    def create_chart(self, sheet_name, chart_type, data_range, position):
        try:
            if not self.current_wb:
                self.current_wb = load_workbook(self.file_path)
                
            ws = self.current_wb[sheet_name]
            chart_map = {"bar": BarChart, "line": LineChart, "pie": PieChart}
            chart = chart_map.get(chart_type, BarChart)()

            data = Reference(ws, min_col=data_range[0], min_row=data_range[1], max_col=data_range[2], max_row=data_range[3])
            chart.add_data(data, titles_from_data=True)
            ws.add_chart(chart, position)

            self.current_wb.save(self.file_path)
        except Exception as e:
            print(f"Error creating chart: {str(e)}")

    def create_pivot(self, sheet_name, data, pivot_index, pivot_columns, pivot_values, aggfunc="sum", start_row=0, start_col=0):
        try:
            if not self.current_wb:
                self.current_wb = load_workbook(self.file_path)
                
            ws = self.current_wb[sheet_name]
            df = pd.DataFrame(data)
            pivot_df = pd.pivot_table(df, index=pivot_index, columns=pivot_columns, values=pivot_values, aggfunc=aggfunc, fill_value=0)
            for row in dataframe_to_rows(pivot_df, index=True, header=True):
                ws.append(row)

            self.current_wb.save(self.file_path)
            return f"Pivot table created successfully in sheet {sheet_name}"
        
        except Exception as e:
            return f"Error creating pivot: {str(e)}"

    def set_worksheet(self, sheet_name):
        self.active_ws = sheet_name
